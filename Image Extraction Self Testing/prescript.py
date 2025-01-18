import zipfile
import os
import openpyxl
import nltk
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
from nltk.stem import PorterStemmer
from nltk import pos_tag
from collections import Counter
import matplotlib.pyplot as plt
from openpyxl import Workbook

# Download NLTK resources
nltk.download('stopwords')
nltk.download('punkt')
nltk.download('averaged_perceptron_tagger')

# Define the path to the Excel file
excel_file = 'my.xlsx'

# Create a directory to save extracted images
output_dir = 'extracted_images'
os.makedirs(output_dir, exist_ok=True)

# Function to process tags using NLP
def process_tag(tag):
    tag = tag.lower()
    words = word_tokenize(tag)
    stop_words = set(stopwords.words('english'))
    filtered_words = [word for word in words if word not in stop_words]
    pos_tags = pos_tag(filtered_words)
    stemmer = PorterStemmer()
    stemmed_words = [(stemmer.stem(word), pos) for word, pos in pos_tags]
    return stemmed_words

# Extract images from the Excel file
with zipfile.ZipFile(excel_file, 'r') as zipf:
    for file_info in zipf.infolist():
        if file_info.filename.startswith('xl/media/') and file_info.filename.endswith(('png', 'jpg', 'jpeg')):
            image_name = os.path.basename(file_info.filename)
            image_path = os.path.join(output_dir, image_name)
            with open(image_path, 'wb') as img_file:
                img_file.write(zipf.read(file_info.filename))

# Load the Excel workbook
workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active

# Rename images based on tags
for img_file in os.listdir(output_dir):
    img_path = os.path.join(output_dir, img_file)
    row_number = int(''.join(filter(str.isdigit, img_file)))
    adjusted_row = row_number + 1  # Adjust row if headings exist
    tag = sheet[f'B{adjusted_row}'].value

    if tag:
        new_path = os.path.join(output_dir, f"{tag.replace(' ', '_')}.png")
        counter = 1
        while os.path.exists(new_path):
            new_path = os.path.join(output_dir, f"{tag.replace(' ', '_')}_{counter}.png")
            counter += 1
        os.rename(img_path, new_path)

# Create a new workbook for processed tags
processed_wb = Workbook()
processed_sheet = processed_wb.active
processed_sheet.title = "Processed Tags"

# Write headers to the new Excel file
processed_sheet.append(["Original Tag", "Processed Tag"])

# Collect words for visualization
all_words = []

# Process each tag and save it to the new Excel file
for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
    original_tag = row[1]
    if original_tag:
        processed_tag = process_tag(original_tag)
        processed_tag_str = " ".join([f"{word}/{pos}" for word, pos in processed_tag])
        processed_sheet.append([original_tag, processed_tag_str])
        all_words.extend([word for word, pos in processed_tag])

# Save the processed tags to a new Excel file
output_excel_file = os.path.join(output_dir, "processed_tags.xlsx")
try:
    processed_wb.save(output_excel_file)
    print(f"Processed tags saved to '{output_excel_file}'.")
except Exception as e:
    print(f"Error saving processed tags: {e}")

# Generate word frequencies
word_frequencies = Counter(all_words)

# Display the word frequencies as a bar chart
plt.figure(figsize=(10, 5))
plt.bar(word_frequencies.keys(), word_frequencies.values())
plt.title("Word Frequencies from Tags")
plt.xticks(rotation=45)
plt.show()

# Print completion message
print("Images extracted, renamed, and tags processed successfully!")